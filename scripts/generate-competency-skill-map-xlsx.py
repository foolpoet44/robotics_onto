#!/usr/bin/env python3

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = (
    Path.home() / ".agent" / "diagrams" / "직무역량_스킬_매핑표_2026.xlsx"
)


def load_mapping_payload():
    """HTML 생성 스크립트의 매핑 결과를 그대로 재사용해 산출물 간 기준 차이를 막는다."""
    script = (
        "const m=require('./scripts/generate-competency-skill-map-html.js');"
        "console.log(JSON.stringify({"
        "assessment:m.competencyDataset.assessment,"
        "source:m.competencyDataset.source,"
        "summary:m.competencyDataset.summary,"
        "colleges:m.competencyDataset.colleges,"
        "mappedRows:m.mappedRows,"
        "byCollege:m.byCollege,"
        "domainCounts:m.domainCounts"
        "}));"
    )
    completed = subprocess.run(
        ["node", "-e", script],
        cwd=REPO_ROOT,
        check=True,
        text=True,
        capture_output=True,
    )
    return json.loads(completed.stdout)


def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8DEE9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = max([len(header) * 1.8, *(min(len(value), 42) for value in values)])
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 52)


def write_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def build_summary_rows(payload):
    rows = [
        ["원천 문서", payload["source"]["title"]],
        ["시트", payload["source"]["sheetName"]],
        ["평가 차수", payload["assessment"]["roundName"]],
        ["평가 인원", payload["summary"]["employeeCount"]],
        ["평가 역량 행", payload["summary"]["competencyCount"]],
        ["고유 직무역량(대/중/소)", len(payload["mappedRows"])],
        [
            "연결된 프로젝트 스킬",
            len({skill["skill_id"] for row in payload["mappedRows"] for skill in row["skills"]}),
        ],
    ]
    rows.append(["", ""])
    rows.append(["전문영역", "직무역량 수 / 평가 인원 / 연결 스킬"])

    for area in payload["byCollege"]:
        rows.append(
            [
                area["nameKo"],
                f'{area["competencyCount"]}개 / {area["employeeCount"]}명 / {area["skillCount"]}개',
            ]
        )
    return rows


def build_grouped_rows(mapped_rows):
    rows = []
    for row in mapped_rows:
        skill_ids = [skill["skill_id"] for skill in row["skills"]]
        skill_names = [skill["preferred_label_ko"] for skill in row["skills"]]
        skill_domains = sorted({skill["domain_en"] for skill in row["skills"]})
        skill_levels = sorted({str(skill["proficiency_level"]) for skill in row["skills"]})

        rows.append(
            [
                row["major"],
                row["middle"],
                row["minor"],
                f'{row["major"]} > {row["middle"]} > {row["minor"]}',
                row["collegeNameKo"],
                row["count"],
                row["averageScore"],
                row["confidence"],
                len(row["skills"]),
                "\n".join(skill_ids),
                "\n".join(skill_names),
                "\n".join(skill_domains),
                ", ".join(skill_levels),
                row["note"],
            ]
        )
    return rows


def build_expanded_rows(mapped_rows):
    rows = []
    for row in mapped_rows:
        # 엑셀에서 검토자가 스킬 하나씩 승인/수정할 수 있도록 후보를 펼친다.
        for index, skill in enumerate(row["skills"], start=1):
            rows.append(
                [
                    row["major"],
                    row["middle"],
                    row["minor"],
                    row["collegeNameKo"],
                    row["count"],
                    row["averageScore"],
                    row["confidence"],
                    index,
                    skill["skill_id"],
                    skill["preferred_label_ko"],
                    skill["preferred_label_en"],
                    skill["domain"],
                    skill["domain_en"],
                    skill["proficiency_level"],
                    skill["skill_type"],
                    skill.get("smartfactory_context", ""),
                    row["note"],
                    "",
                ]
            )
    return rows


def main():
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)

    payload = load_mapping_payload()
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "요약"
    write_rows(ws_summary, ["항목", "값"], build_summary_rows(payload))
    style_sheet(ws_summary)

    ws_grouped = wb.create_sheet("매핑표_대중소기준")
    write_rows(
        ws_grouped,
        [
            "기술역량 대분류",
            "기술역량 중분류",
            "기술역량 소분류",
            "직무역량 경로",
            "전문영역",
            "평가 인원",
            "평균 점수",
            "매핑 신뢰도",
            "후보 스킬 수",
            "후보 스킬 ID",
            "후보 스킬명",
            "후보 스킬 도메인",
            "후보 레벨",
            "매핑 근거",
        ],
        build_grouped_rows(payload["mappedRows"]),
    )
    add_table(ws_grouped, "CompetencySkillMapGrouped")

    ws_expanded = wb.create_sheet("매핑원장_스킬펼침")
    write_rows(
        ws_expanded,
        [
            "기술역량 대분류",
            "기술역량 중분류",
            "기술역량 소분류",
            "전문영역",
            "평가 인원",
            "평균 점수",
            "매핑 신뢰도",
            "후보 순번",
            "스킬 ID",
            "스킬명",
            "스킬명 EN",
            "스킬 도메인",
            "스킬 도메인 EN",
            "숙련 레벨",
            "스킬 유형",
            "스마트팩토리 맥락",
            "매핑 근거",
            "검토 의견",
        ],
        build_expanded_rows(payload["mappedRows"]),
    )
    add_table(ws_expanded, "CompetencySkillMapExpanded")

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
